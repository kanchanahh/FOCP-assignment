import tkinter as tk
from tkinter import ttk, messagebox
from tkinter import PhotoImage
import openpyxl
from openpyxl import Workbook
import os

FILE_NAME = "futsal_bookings.xlsx"
User_File = "users.xlsx"
Matches = "matches.xlsx"
player = "player.xlsx"

if not os.path.exists(User_File):
    wb = Workbook()
    ws = wb.active
    ws.append(["Username", "Email", "Password"])  
    wb.save(User_File)

# ---register function---
def register_user(username, email, password):
    wb = openpyxl.load_workbook(User_File)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == username:
            messagebox.showerror("Error", "Username already exists!")
            return
    sheet.append([username, email, password])
    wb.save(User_File)
    messagebox.showinfo("Success", "Registration successful! Please log in.")

def clear_register_fields():
    username_register.delete(0, tk.END)
    email_register.delete(0, tk.END)
    password_register.delete(0, tk.END)

# ---login function---
def login_user(username, password):
    wb = openpyxl.load_workbook(User_File)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == username and row[2] == password:
            show_home_page()
            return
    messagebox.showerror("Error", "Invalid username or password!")

def show_login_page():
    register_frame.pack_forget()
    home_frame.pack_forget()
    login_frame.pack(fill="both", expand=True)

def show_register_page():
    login_frame.pack_forget()
    register_frame.pack(fill="both", expand=True)    

def show_home_page():
    login_frame.pack_forget()
    register_frame.pack_forget()
    home_frame.pack(fill="both", expand=True)

# ---add frame function---
def add():
    home_frame.pack_forget()
    add_frame.pack(fill="both", expand=True)  

def add_player():
    name = entry_player_name.get().strip()
    contact = entry_player_contact.get().strip()
    team = entry_team.get().strip() 

    if not name or not contact or not team:
        messagebox.showwarning("Warning", "Please fill out all fields.")
        return

    try:
        if not os.path.exists(player):
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Players"
            worksheet.append(["Name", "Contact", "Team"])
        else:
            workbook = openpyxl.load_workbook(player)
            worksheet = workbook.active

        worksheet.append([name, contact, team])
        workbook.save(player)

        messagebox.showinfo("Success", "Player details saved successfully!")
        clear_add_inputs()
    except Exception as e:
        messagebox.showerror("Error", f"Failed to save player data: {e}")

def load_players():
    if not os.path.exists(player):
        messagebox.showinfo("Info", "No player data found!")
        return

    for row in player_table.get_children():
        player_table.delete(row)

    try:
        workbook = openpyxl.load_workbook(player)
        worksheet = workbook.active

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            player_table.insert("", tk.END, values=row)
    except Exception as e:
        messagebox.showerror("Error", f"Failed to load player data: {e}")

def clear_add_inputs():
    entry_name.delete(0, tk.END)
    entry_contact.delete(0, tk.END)
    entry_team.delete(0, tk.END)

def go_back_from_add():
    add_frame.pack_forget()
    home_frame.pack(fill="both", expand=True)

#---match frame function--- 
def show_match_viewer():
    home_frame.pack_forget()
    match_viewer_frame.pack(fill="both", expand=True) 

def save_match(team1, team2, field_entry, time_entry):
    team1_name = team1.get().strip()
    team2_name = team2.get().strip()
    field_name = field_entry.get().strip()
    match_time = time_entry.get().strip()

    print(f"Name1: {team1_name}, Name2: {team2_name}, field_name: {field_name}, match_time: {match_time}")

    if not team1_name or not team2_name or not field_name or not match_time:
        messagebox.showerror("Error", "Please fill out all fields!")
        return

    if not os.path.exists(Matches):
        wb = Workbook()
        ws = wb.active
        ws.title = "Matches"
        ws.append(["Team 1", "Team 2", "Field", "Time"]) 
        wb.save(Matches)

    wb = openpyxl.load_workbook(Matches)
    ws = wb.active
    ws.append([team1_name, team2_name, field_name, match_time])
    wb.save(Matches)

    messagebox.showinfo("Success", "Match data saved successfully!")
    clear_match_fields(team1, team2, field_entry, time_entry)

def clear_match_fields(team1, team2, field_entry, time_entry):
    team1.delete(0, tk.END)
    team2.delete(0, tk.END)
    field_entry.delete(0, tk.END)
    time_entry.delete(0, tk.END) 

def show_matches(table):
    try:
        if not os.path.exists(Matches):  
            messagebox.showerror("Error", "No match data found!")
            return

        for row in table.get_children():
            table.delete(row)

        wb = openpyxl.load_workbook(Matches)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            team = f"{row[0]} vs {row[1]}"
            field = row[2]
            time = row[3]
            table.insert("", "end", values=(team, field, time))

    except Exception as e:
        messagebox.showerror("Error", f"Failed to load matches: {e}")

def go_back_from_match():
    match_viewer_frame.pack_forget()
    home_frame.pack(fill="both", expand=True) 

# ---booking frame function---
def booking():
    home_frame.pack_forget()
    booking_frame.pack(fill="both", expand=True)

def initialize_excel():
    try:
        workbook = openpyxl.load_workbook(FILE_NAME)
        worksheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["Name", "Contact", "Date", "Time", "Field"])
        workbook.save(FILE_NAME)

def save_booking():
    name = entry_name.get().strip()
    contact = entry_contact.get().strip()
    date = entry_date.get().strip()
    time = entry_time.get().strip()
    field = entry_field.get().strip()

    if not name or not contact or not date or not time or not field:
        messagebox.showerror("Error", "All fields are required!")
        return

    if not contact.isdigit():
        messagebox.showerror("Error", "Contact must be numeric!")
        return

    for row in table.get_children():
        existing_date, existing_time, existing_field = table.item(row, 'values')[2:5]
        if date == existing_date and time == existing_time and field == existing_field:
            messagebox.showerror("Error", "This time slot is already booked!")
            return

    table.insert("", tk.END, values=(name, contact, date, time, field))

    workbook = openpyxl.load_workbook(FILE_NAME)
    worksheet = workbook.active
    worksheet.append([name, contact, date, time, field])
    workbook.save(FILE_NAME)

    messagebox.showinfo("Success", "Booking saved successfully!")
    clear_all()

def load_bookings():
    workbook = openpyxl.load_workbook(FILE_NAME)
    worksheet = workbook.active
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        table.insert("", tk.END, values=row)

def clear_all():
    entry_name.delete(0, tk.END)
    entry_contact.delete(0, tk.END)
    entry_date.delete(0, tk.END)
    entry_time.delete(0, tk.END)
    entry_field.delete(0, tk.END)

def delete_booking():
    selected_item = table.selection()
    if not selected_item:
        messagebox.showerror("Error", "No booking selected!")
        return

    for item in selected_item:
        table.delete(item)

    workbook = openpyxl.load_workbook(FILE_NAME)
    worksheet = workbook.active
    worksheet.delete_rows(2, worksheet.max_row)  

  
    worksheet.append(["Name", "Contact", "Date", "Time", "Field"])

    for row in table.get_children():
        worksheet.append(table.item(row, "values"))

    workbook.save(FILE_NAME)
    messagebox.showinfo("Success", "Booking deleted successfully!")

def go_back_from_booking():
    booking_frame.pack_forget()
    home_frame.pack(fill="both", expand=True)     

def exit_app():
    root.quit()
 
# Main application window
root = tk.Tk()
root.title("Futsal Booking System")
root.geometry("600x400")
root.configure(bg="#815649")

# Load image
logo_image = PhotoImage(file="football.png")
logo_image = logo_image.subsample(3, 3)

# ---Login Frame---
login_frame = tk.Frame(root, bg="#815649")
login_frame.pack(fill="both", expand=True)

tk.Label(login_frame, text="Login", font=("Arial", 16), bg="#815649", fg="white").pack(pady=10)
tk.Label(login_frame, text="Username:", bg="#815649", fg="white").pack(pady=5)
username_login = tk.Entry(login_frame)
username_login.pack(pady=5)
tk.Label(login_frame, text="Password:", bg="#815649", fg="white").pack(pady=5)
password_login = tk.Entry(login_frame, show="*")
password_login.pack(pady=5)

login_button = tk.Button(login_frame, text="Login", bg="#815649", fg="white",
                         command=lambda: login_user(username_login.get(), password_login.get()))
login_button.pack(pady=10)
register_button = tk.Button(login_frame, text="Register", bg="#815649", fg="white", command=show_register_page)
register_button.pack()

# ---Registration Frame---
register_frame = tk.Frame(root, bg="#815649")
tk.Label(register_frame, text="Register", font=("Arial", 16), bg="#815649", fg="white").pack(pady=10)
tk.Label(register_frame, text="Username:", bg="#815649", fg="white").pack(pady=5)
username_register = tk.Entry(register_frame)
username_register.pack(pady=5)
tk.Label(register_frame, text="Email:", bg="#815649", fg="white").pack(pady=5)
email_register = tk.Entry(register_frame)
email_register.pack(pady=5)
tk.Label(register_frame, text="Password:", bg="#815649", fg="white").pack(pady=5)
password_register = tk.Entry(register_frame, show="*")
password_register.pack(pady=5)

register_button = tk.Button(register_frame, text="Register", bg="#815649", fg="white",
                            command=lambda: register_user(username_register.get(), email_register.get(), password_register.get()))
register_button.pack(pady=10)
clear_button = tk.Button(register_frame, text="Clear All", bg="#815649", fg="white", command=clear_register_fields)
clear_button.pack(pady=10)
back_to_login_button = tk.Button(register_frame, text="Back to Login", bg="#815649", fg="white", command=show_login_page)
back_to_login_button.pack()

# ---Home Page Frame---
home_frame = tk.Frame(root, bg="#815649")
header_frame_home = tk.Frame(home_frame, bg="#815649")
header_frame_home.pack(fill="x", pady=10)

logo_home = tk.Label(header_frame_home, image=logo_image, bg="#815649")
logo_home.pack(side="top", padx=10, pady=5)

name_label_home = tk.Label(header_frame_home, text="Futsal Booking System", font=("Arial", 16), bg="#815649", fg="white")
name_label_home.pack(side="top", padx=10, pady=5)

content_frame_home = tk.Frame(home_frame, bg="#FAEBD7", width=400, height=300)
content_frame_home.pack(padx=20, pady=10, fill="both", expand=True)

add_button = tk.Button(content_frame_home, text="Add", bg="#815649", fg="white", font=("Arial", 12),command=add, width=20)
add_button.pack(pady=10)

book_button = tk.Button(content_frame_home, text="Book futsal ground", bg="#815649", fg="white", font=("Arial", 12), command=booking, width=20)
book_button.pack(pady=10)

view_button = tk.Button(content_frame_home, text="View matches", bg="#815649", fg="white", font=("Arial", 12), command=show_match_viewer, width=20)
view_button.pack(pady=10)

exit_button_home = tk.Button(content_frame_home, text="Exit", font=("Arial", 12), bg="#815649", fg="white", command=root.quit)
exit_button_home.pack(side="bottom", anchor="se", padx=10, pady=10)

# Start with login page
show_login_page()

# ---Add Frame---
add_frame = tk.Frame(root, bg="#815649")

frame_header_add = tk.Frame(add_frame, bg="#815649")
frame_header_add.pack(pady=10, anchor="w")
logo_add = tk.Label(frame_header_add, image=logo_image, bg="#815649")
logo_add.pack(side="left", padx=10)

frame_inputs_add = tk.Frame(add_frame, bg="#815649")
frame_inputs_add.pack(pady=10)

tk.Label(frame_inputs_add, text="Name", bg="#815649", fg="white").grid(row=0, column=0, padx=5, pady=5, sticky="e")
entry_player_name = tk.Entry(frame_inputs_add, width=30)
entry_player_name.grid(row=0, column=1, padx=5, pady=5)

tk.Label(frame_inputs_add, text="Contact", bg="#815649", fg="white").grid(row=1, column=0, padx=5, pady=5, sticky="e")
entry_player_contact = tk.Entry(frame_inputs_add, width=30)
entry_player_contact.grid(row=1, column=1, padx=5, pady=5)

tk.Label(frame_inputs_add, text="Team", bg="#815649", fg="white").grid(row=2, column=0, padx=5, pady=5, sticky="e")
entry_team = tk.Entry(frame_inputs_add, width=30)
entry_team.grid(row=2, column=1, padx=5, pady=5)

save_button_add = tk.Button(frame_inputs_add, text="Save", command=add_player, bg="#FAEBD7")
save_button_add.grid(row=3, column=0, pady=10)

save_button_add = tk.Button(frame_inputs_add, text="Save", command=add_player, bg="#FAEBD7")

clear_button_add = tk.Button(frame_inputs_add, text="Clear Input Fields", command=lambda: clear_add_inputs(), bg="#FAEBD7")
clear_button_add.grid(row=3, column=1, pady=10)

frame_table_add = tk.Frame(add_frame, bg="#815649")
frame_table_add.pack(pady=10)

player_table = ttk.Treeview(frame_table_add, columns=("Name", "Contact", "Team"), show="headings", style="Treeview")
player_table.heading("Name", text="Name")
player_table.heading("Contact", text="Contact")
player_table.heading("Team", text="Team")
player_table.column("Name", anchor="center", width=150)
player_table.column("Contact", anchor="center", width=150)
player_table.column("Team", anchor="center", width=150)
player_table.pack(pady=10, padx=10, fill="both", expand=True)

load_button_add = tk.Button(frame_table_add, text="Load Players", command=load_players, bg="#FAEBD7")
load_button_add.pack(side="left", padx=10, pady=10)

back_button_add = tk.Button(frame_table_add, text="Back", command=go_back_from_add, bg="#815649", fg="white", font=("Arial", 12))
back_button_add.pack(side="right", padx=10, pady=10)

# ---Match Viewer Frame---
match_viewer_frame = tk.Frame(root, bg="#815649")

frame_header_match = tk.Frame(match_viewer_frame, bg="#815649")
frame_header_match.pack(pady=10, anchor="w")

logo_add = tk.Label(frame_header_match, image=logo_image, bg="#815649")
logo_add.pack(side="left", padx=10)

frame_inputs_match = tk.Frame(match_viewer_frame, bg="#815649")
frame_inputs_match.pack(pady=10)

tk.Label(frame_inputs_match, text="Team 1", bg="#815649", fg="white").grid(row=0, column=0)
entry_team1 = tk.Entry(frame_inputs_match)
entry_team1.grid(row=0, column=1)

tk.Label(frame_inputs_match, text="Team 2", bg="#815649", fg="white").grid(row=1, column=0)
entry_team2 = tk.Entry(frame_inputs_match)
entry_team2.grid(row=1, column=1)

tk.Label(frame_inputs_match, text="Field", bg="#815649", fg="white").grid(row=2, column=0)
match_entry_field = tk.Entry(frame_inputs_match)
match_entry_field.grid(row=2, column=1)

tk.Label(frame_inputs_match, text="Time", bg="#815649",fg="white").grid(row=3, column=0)
match_entry_time = tk.Entry(frame_inputs_match)
match_entry_time.grid(row=3, column=1)

save_button_match = tk.Button(frame_inputs_match, text="Save Match", command=lambda: save_match(entry_team1, entry_team2, match_entry_field, match_entry_time), bg="#faebd7")
save_button_match.grid(row=4, column=0, pady=10)

clear_button_match = tk.Button(frame_inputs_match, text="Clear Input Fields", command=lambda: clear_match_fields(entry_team1, entry_team2, match_entry_field, match_entry_time), bg="#faebd7")
clear_button_match.grid(row=4, column=1)
# Table to display matches
frame_table = tk.Frame(match_viewer_frame, bg="#815649")
frame_table.pack(pady=10)

columns = ("team", "field", "time")
match_table = ttk.Treeview(frame_table, columns=columns, show="headings", height=8)
match_table.heading("team", text="Team")
match_table.heading("field", text="Field")
match_table.heading("time", text="Time")
match_table.pack()

show_button_match = tk.Button(frame_table,text="Show Matches",command=lambda: show_matches(match_table),bg="#faebd7",)
show_button_match.pack(side="left", pady=5)
back_button_match = tk.Button(frame_table,text="Back",  command=go_back_from_match, bg="#815649",fg="white",font=("Arial", 12),)
back_button_match.pack(side="right", padx=10, pady=10)

#---Booking Page Frame---
booking_frame = tk.Frame(root, bg="#815649") 

frame_header = tk.Frame(booking_frame, bg="#815649")
frame_header.pack(pady=10, anchor="w")
logo = tk.Label(frame_header, image=logo_image, bg="#815649")
logo.pack(side="left", padx=10)

frame_inputs = tk.Frame(booking_frame, bg="#815649")
frame_inputs.pack(pady=10)
tk.Label(frame_inputs, text="Name", bg="#815649", fg="white").grid(row=0, column=0)
entry_name = tk.Entry(frame_inputs)
entry_name.grid(row=0, column=1)
tk.Label(frame_inputs, text="Contact", bg="#815649", fg="white").grid(row=1, column=0)
entry_contact = tk.Entry(frame_inputs)
entry_contact.grid(row=1, column=1)
tk.Label(frame_inputs, text="Date (YYYY-MM-DD)", bg="#815649", fg="white").grid(row=2, column=0)
entry_date = tk.Entry(frame_inputs)
entry_date.grid(row=2, column=1)
tk.Label(frame_inputs, text="Time (hr:min)", bg="#815649", fg="white").grid(row=3, column=0)
entry_time = tk.Entry(frame_inputs)
entry_time.grid(row=3, column=1)
tk.Label(frame_inputs, text="Ground", bg="#815649", fg="white").grid(row=4, column=0)
entry_field = tk.Entry(frame_inputs)
entry_field.grid(row=4, column=1)
tk.Button(frame_inputs, text="Save Booking", command=save_booking, bg="#faebd7").grid(row=5, column=0, pady=10)
tk.Button(frame_inputs, text="Clear Input Fields", command=clear_all, bg="#faebd7").grid(row=5, column=1)

frame_table = tk.Frame(booking_frame, bg="#815649")
frame_table.pack()
table = ttk.Treeview(frame_table, columns=("Name", "Contact", "Date", "Time", "Field"), show="headings", style="Treeview")
table.heading("Name", text="Name")
table.heading("Contact", text="Contact")
table.heading("Date", text="Date")
table.heading("Time", text="Time")
table.heading("Field", text="Field")
table.pack()
delete_button=tk.Button(frame_table, text="Delete Booking", command=delete_booking, bg="#faebd7")
delete_button.pack(side="left",pady=10)

back_button_booking = tk.Button(frame_table, text="Back", command=go_back_from_booking, bg="#815649", fg="white", font=("Arial", 12))
back_button_booking.pack(side="right",pady=10)

root.mainloop()